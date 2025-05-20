from typing import Any
from datetime import datetime

from django.contrib import admin, messages
from rangefilter.filters import  DateRangeFilterBuilder
from django.db.models.query import QuerySet
from django.http import HttpRequest, HttpResponse, QueryDict
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from utils.django.model_admin import ModelAdmin 

from .models import Company, ContactRequest


@admin.register(Company)
class CompanyAdmin(ModelAdmin): ...

@admin.register(ContactRequest)
class ContactRequestAdmin(ModelAdmin):
    list_display = ("fullname", "email", "phone", "company", "subject", "created_at" )
    list_filter = (
        ("created_at", DateRangeFilterBuilder()),
    )
    actions = ["download_report",]

    def has_add_permission(self, request: HttpRequest) -> bool:
        return False

    def has_change_permission(self, request: HttpRequest, obj: Any | None = ...) -> bool:
        return False

    def has_delete_permission(self, request: HttpRequest, obj: Any | None = ...) -> bool:
        return False


    @admin.action(description="Descargar Reporte")
    def download_report(self, request:HttpRequest, queryset:QuerySet[ContactRequest]):
        queryset_list = list(queryset)
        if not queryset_list:
            self.message_user(request, "No hay registros en el rango de fechas seleccionado o en la selección actual.", level=messages.WARNING)
            return None
        headers = ["Full Name", "Email", "Phone", "Company", "Subject", "Created At"]
        return generate_excel_report(queryset_list, headers)

    

def generate_excel_report(queryset: list, headers: list[str]) -> HttpResponse:
    """
    Genera un archivo Excel a partir de un queryset y una lista de encabezados.
    Devuelve una respuesta HTTP con el archivo generado.

    :param queryset: Lista de objetos que se convertirán en filas en el archivo Excel.
    :param headers: Lista de encabezados para las columnas del archivo Excel.
    :return: HttpResponse con el archivo Excel.
    """
    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise ValueError("Error al crear la hoja de Excel.")

    # Agregar encabezados
    ws.append(headers)

    # Llenar las filas con los datos
    for obj in queryset:
        ws.append([
            obj.fullname,
            obj.email,
            obj.phone,
            obj.company,
            obj.subject,
            obj.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    # Ajustar el ancho de las columnas
    for i, col in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max(12, max_length + 2)

    # Crear la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=contact_requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response